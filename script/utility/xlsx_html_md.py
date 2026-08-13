# @audit this are a temporary solution for an issue below 
# [lynnhubbz/Astrocosmos#25](https://github.com/lynnhubbz/Astrocosmos/issues/25)
#  
# this python script are brainstormed and will be reviewed later.
# The main objective of this script is to convert an excel file to html file then to markdown.
# it is also saved in html so it can be version controled.
#
# The main flow of this script are:
# 1. read the list of watched files
# 2. turns excel into pandas dataframe
# 3. save it into a template html
# 4. embed the html into markdown


import os
import time
import yaml
import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# --- CONFIGURATION PATH ---
CONFIG_FILE_PATH = "script/utility/watchlist_xlsx.yaml"


def load_pipeline_configuration():
    """
    Reads the batch configuration YAML file and extracts global settings and jobs.
    Supports the updated 'jobs' key structure seamlessly.
    """
    if not os.path.exists(CONFIG_FILE_PATH):
        return {}, []

    with open(CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
        try:
            config = yaml.safe_load(f) or {}
        except Exception as e:
            print(f"Error parsing YAML configuration file: {e}")
            return {}, []

    global_settings = config.get("global_settings", {})
    # Explicitly map the new 'jobs' key block
    jobs = config.get("jobs", [])
    return global_settings, jobs

def parse_sheet_to_html(ws):
    """
    HEAVY LIFTING WORKER: Converts a single sheet object (ws) into an HTML table.
    This contains your original openpyxl merged-cell parsing logic.
    """
    top_left_cells = {}
    hidden_cells = set()
    
    # 1. Map out Excel's native cell merge structural boundaries
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        top_left_cells[(min_row, min_col)] = (rowspan, colspan)
        
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                hidden_cells.add((r, c))
                
    # 2. Rebuild structural HTML lines cleanly cell by cell
    html_lines = ['<table class="tidy-table">']
    for r in range(1, ws.max_row + 1):
        html_lines.append('  <tr>')
        for c in range(1, ws.max_column + 1):
            if (r, c) in hidden_cells:
                continue
                
            cell = ws.cell(row=r, column=c)
            val = cell.value if cell.value is not None else ""
            
            attrs = []
            if (r, c) in top_left_cells:
                rowspan, colspan = top_left_cells[(r, c)]
                if rowspan > 1: attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1: attrs.append(f'colspan="{colspan}"')
                    
            attr_str = " " + " ".join(attrs) if attrs else ""
            tag = 'th' if r == 1 else 'td'
            
            escaped_val = str(val).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            html_lines.append(f'    <{tag}{attr_str}>{escaped_val}</{tag}>')
        html_lines.append('  </tr>')
        
    html_lines.append('</table>')
    return "\n".join(html_lines)


def run_single_job_pipeline(job, global_settings):
    """
    CONTROLLER: Loops through all tabs, stacks them with ### Headings, 
    and inserts the complete multi-table block into your templates.
    """
    source_excel = job.get("source_excel")
    if not source_excel:
        print(" [!] Execution Error: 'source_excel' path key is REQUIRED. Skipping.")
        return

    if not os.path.exists(source_excel):
        print(f" -> Execution Error: Source Excel file missing: '{source_excel}'")
        return

    excel_dir = os.path.dirname(source_excel) or "."
    excel_base_name = os.path.splitext(os.path.basename(source_excel))[0]

    output_markdown = job.get("output_markdown")
    if not output_markdown:
        output_markdown = os.path.join(excel_dir, f"{excel_base_name}.md")

    resolved_title = job.get("title") or os.path.splitext(os.path.basename(output_markdown))[0]
    html_template_file = job.get("html_template", global_settings.get("html_template", "html_template.html"))
    markdown_template_file = job.get("markdown_template", global_settings.get("markdown_template", "template.md"))
    placeholder = job.get("placeholder", global_settings.get("placeholder", "<!-- TABLE_PLACEHOLDER -->"))
    output_html_cfg = job.get("output_html")

    print(f" -> Processing Multi-Tab File: '{source_excel}'")
    
    try:
        # Load the workbook framework
        wb = openpyxl.load_workbook(source_excel, data_only=True)
        
        # Verify templates exist before running the loops
        if not os.path.exists(html_template_file) or not os.path.exists(markdown_template_file):
            print("    [!] Structural Error: HTML or Markdown template path is missing.")
            return

        with open(html_template_file, "r", encoding="utf-8") as html_temp:
            html_layout_blueprint = html_temp.read()

        # Storage lists for the compiled tabs
        compiled_html_blocks = []
        compiled_markdown_blocks = []

        # --- NEW CODE: LOOP THROUGH ALL WORKBOOK TABS ---
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Skip empty tabs completely so your output stays tidy
            if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(row=1, column=1).value is None:
                continue
                
            # 1. Generate the isolated HTML <table> element block
            html_table_element = parse_sheet_to_html(ws)
            
            # 2. Wrap it inside your specific custom container layout blueprint
            fully_wrapped_html = html_layout_blueprint.replace(placeholder, html_table_element)
            
            # Sanitize the sheet name to make a clean layout hook CSS class (e.g., "Sheet 1" -> "sheet_1")
            safe_css_class = sheet_name.lower().replace(" ", "_")
            
            # 3. Create the HTML compilation chunk
            html_chunk = f'<!-- START SECTION: {sheet_name} -->\n<div class="tab-container tab-section-{safe_css_class}">\n  {fully_wrapped_html}\n</div>\n<!-- END SECTION -->\n'
            compiled_html_blocks.append(html_chunk)
            
            # 4. Create the Markdown compilation chunk with your requested Level 3 heading separator
            md_chunk = f'### {sheet_name}\n\n<div class="tab-container tab-section-{safe_css_class}">\n  {fully_wrapped_html}\n</div>\n'
            compiled_markdown_blocks.append(md_chunk)

        if not compiled_markdown_blocks:
            print("    -> Warning: All sheets were completely empty. Output skipped.")
            return

        # --- SAVE THE COMBINED STANDALONE HTML FILE ---
        html_out_path = None
        if output_html_cfg is None or (isinstance(output_html_cfg, bool) and output_html_cfg):
            html_out_path = os.path.join(excel_dir, f"{excel_base_name}.html")
        elif isinstance(output_html_cfg, str) and output_html_cfg.strip():
            html_out_path = output_html_cfg

        if html_out_path:
            with open(html_out_path, "w", encoding="utf-8") as html_out:
                # Merge all individual wrapped HTML tables into one document block string
                html_out.write("\n<br>\n".join(compiled_html_blocks))
            print(f"    -> Mirror Generated successfully -> '{html_out_path}'")

        # --- SAVE THE FINAL DOCUMENTATION MARKDOWN FILE ---
        with open(markdown_template_file, "r", encoding="utf-8") as md_temp:
            markdown_blueprint = md_temp.read()

        # Combine all Markdown/Heading blocks together
        all_tables_markdown_string = "\n".join(compiled_markdown_blocks)
        
        # Inject everything into the parent Markdown documentation file template
        final_markdown_payload = markdown_blueprint.replace(placeholder, all_tables_markdown_string)
        final_markdown_payload = final_markdown_payload.replace("title: {{TITLE}}", f"title: {resolved_title}")

        output_dir = os.path.dirname(output_markdown)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        with open(output_markdown, "w", encoding="utf-8") as production_markdown:
            production_markdown.write(final_markdown_payload)
        print(f"    -> Production Build Complete -> '{output_markdown}'")

    except PermissionError:
        print(f"    [!] Access Denied: File locked inside Excel. It will catch changes on the next save cycle.")
    except Exception as e:
        print(f"    [!] Conversion engine exception: {e}")


def execute_entire_batch_pipeline():
    """ Runs through every conversion assignment configured in the YAML manifest file sequentially. """
    print(f"[{time.strftime('%H:%M:%S')}] Launching comprehensive batch conversion tasks...")
    global_settings, jobs = load_pipeline_configuration()
    
    if not jobs:
        print(" -> Notice: No file track conversion assignments registered inside configuration map.")
        return

    for job in jobs:
        run_single_job_pipeline(job, global_settings)
    print(f"[{time.strftime('%H:%M:%S')}] Active batch processing complete.\n")


class BulkExcelStorageChangeHandler(FileSystemEventHandler):
    """ Listens to filesystem modifications to trigger real-time compilation loops selectively. """
    def on_modified(self, event):
        if event.is_directory:
            return
            
        global_settings, jobs = load_pipeline_configuration()
        modified_abs_path = os.path.abspath(event.src_path)

        # Pause briefly to ensure active write cycles clear fully before parsing
        time.sleep(0.5)

        for job in jobs:
            excel_path = job.get("source_excel")
            if excel_path and os.path.abspath(excel_path) == modified_abs_path:
                print(f"[{time.strftime('%H:%M:%S')}] Modification detected for file asset: '{os.path.basename(excel_path)}'")
                run_single_job_pipeline(job, global_settings)
                print("Monitoring folder environments... Click Ctrl+C to terminate runtime loops.\n")


if __name__ == "__main__":
    # Bootstrap configuration manifest parameters file if completely missing
    if not os.path.exists(CONFIG_FILE_PATH):
        with open(CONFIG_FILE_PATH, "w", encoding="utf-8") as f:
            f.write("""global_settings:
  placeholder: "<!-- CONTENT_HERE -->"
  html_template: "test_env/data.html"
  markdown_template: "test_env/data.md"

jobs parameters:
  - title: "Quarterly Analytics Dashboard"
    source_excel: "test_env/metrics.xlsx"
    output_markdown: "test_env/final_report.md"
    output_html: true
""")
        print(f"[*] Instantiated an interactive structural configuration setup script at: '{CONFIG_FILE_PATH}'")

    # Run basic sequential evaluation loops over active registers
    execute_entire_batch_pipeline()

    # Track directory paths dynamically where monitored spreadsheets reside
    global_settings, jobs = load_pipeline_configuration()
    monitored_directories = set()
    for job in jobs:
        excel_path = job.get("source_excel")
        if excel_path:
            dir_path = os.path.dirname(os.path.abspath(excel_path)) or "."
            monitored_directories.add(dir_path)

    if monitored_directories:
        event_handler = BulkExcelStorageChangeHandler()
        observer = Observer()
        
        for watch_dir in monitored_directories:
            if os.path.exists(watch_dir):
                observer.schedule(event_handler, path=watch_dir, recursive=False)
                print(f"Actively tracking workspace storage location directory: '{watch_dir}'")
            
        print("Pipeline core engines monitoring setup online. Press Ctrl+C to stop tracking.\n")
        observer.start()
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\nShutting down pipeline processes safely...")
            observer.stop()
        observer.join()